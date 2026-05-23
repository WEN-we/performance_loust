"""
报告生成服务模块

提供多种格式的测试报告生成能力，包括：
- HTML测试报告（使用jinja2模板，内联CSS/JS，不依赖外部文件）
- PDF报告（使用fpdf2）
- Excel导出（使用openpyxl）
- PNG图表导出（使用matplotlib）
- 报告内容：任务概览、统计数据、响应时间分布、状态码分布
"""

import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from database.db_manager import DatabaseManager
from utils.logger import get_logger
from utils.helpers import ensure_dir, format_duration

logger = get_logger("report_service")


class ReportService:
    """报告生成服务类

    根据任务执行结果生成多种格式的测试报告，
    支持HTML、PDF、Excel和PNG图表。
    """

    def __init__(self, db: DatabaseManager | None = None) -> None:
        """初始化报告生成服务

        Args:
            db: 数据库管理器实例，为None时使用单例
        """
        self._db = db or DatabaseManager()

    def generate_html_report(
        self,
        task_id: int,
        result_id: int | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """生成HTML测试报告

        使用jinja2模板引擎渲染HTML报告，所有CSS和JS内联在HTML中，
        不依赖任何外部文件，可直接在浏览器中打开查看。

        Args:
            task_id: 任务ID
            result_id: 结果记录ID，为None时使用最新结果
            output_path: 输出文件路径，为None时自动生成

        Returns:
            生成的报告文件路径

        Raises:
            ValueError: 任务或结果不存在
        """
        from jinja2 import Template

        report_data = self._collect_report_data(task_id, result_id)

        if output_path is None:
            from config.settings import get_settings
            settings = get_settings()
            export_dir = settings.export_dir
            ensure_dir(export_dir)
            safe_name = report_data["task_info"].get("name", "unnamed").replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = export_dir / f"report_{safe_name}_{timestamp}.html"

        output_path = Path(output_path)
        ensure_dir(output_path.parent)

        template = Template(self._HTML_TEMPLATE)
        html_content = template.render(**report_data)

        output_path.write_text(html_content, encoding="utf-8")
        logger.info("HTML报告已生成，路径=%s", output_path)
        return output_path

    def generate_pdf_report(
        self,
        task_id: int,
        result_id: int | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """生成PDF报告

        使用fpdf2库生成PDF格式的测试报告，
        包含任务概览、统计数据、响应时间分布等核心信息。

        Args:
            task_id: 任务ID
            result_id: 结果记录ID，为None时使用最新结果
            output_path: 输出文件路径，为None时自动生成

        Returns:
            生成的报告文件路径

        Raises:
            ValueError: 任务或结果不存在
        """
        from fpdf import FPDF

        report_data = self._collect_report_data(task_id, result_id)

        if output_path is None:
            from config.settings import get_settings
            settings = get_settings()
            export_dir = settings.export_dir
            ensure_dir(export_dir)
            safe_name = report_data["task_info"].get("name", "unnamed").replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = export_dir / f"report_{safe_name}_{timestamp}.pdf"

        output_path = Path(output_path)
        ensure_dir(output_path.parent)

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        font_paths = [
            Path("C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("C:/Windows/Fonts/simsun.ttc"),
        ]
        chinese_font = None
        for fp in font_paths:
            if fp.exists():
                chinese_font = str(fp)
                break

        if chinese_font:
            pdf.add_font("Chinese", "", chinese_font, uni=True)
            pdf.add_font("Chinese", "B", chinese_font, uni=True)
            title_font = "Chinese"
            body_font = "Chinese"
        else:
            title_font = "Helvetica"
            body_font = "Helvetica"

        pdf.set_font(title_font, "B", 20)
        pdf.cell(0, 15, "Performance Test Report", ln=True, align="C")
        pdf.ln(5)

        pdf.set_font(body_font, "", 10)
        pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
        pdf.ln(10)

        task_info = report_data["task_info"]
        pdf.set_font(title_font, "B", 14)
        pdf.cell(0, 10, "Task Overview", ln=True)
        pdf.set_font(body_font, "", 10)

        overview_items = [
            ("Task Name", str(task_info.get("name", "-"))),
            ("Task Type", str(task_info.get("type", "-"))),
            ("Method", str(task_info.get("method", "-"))),
            ("URL", str(task_info.get("url", "-"))),
            ("Users", str(task_info.get("users", "-"))),
            ("Spawn Rate", str(task_info.get("spawn_rate", "-"))),
            ("Run Time", str(task_info.get("run_time", "-"))),
        ]
        for label, value in overview_items:
            pdf.cell(50, 7, f"  {label}:", border=0)
            pdf.cell(0, 7, value, ln=True)

        pdf.ln(5)

        stats = report_data["stats"]
        pdf.set_font(title_font, "B", 14)
        pdf.cell(0, 10, "Statistics", ln=True)
        pdf.set_font(body_font, "", 10)

        stats_items = [
            ("Total Requests", str(stats.get("total_requests", 0))),
            ("Success Count", str(stats.get("success_count", 0))),
            ("Fail Count", str(stats.get("fail_count", 0))),
            ("Fail Rate", f"{stats.get('fail_rate', 0):.2%}"),
            ("Avg Response Time", f"{stats.get('avg_response_time', 0):.2f} ms"),
            ("Min Response Time", f"{stats.get('min_response_time', 0):.2f} ms"),
            ("Max Response Time", f"{stats.get('max_response_time', 0):.2f} ms"),
            ("P95 Response Time", f"{stats.get('p95_response_time', 0):.2f} ms"),
            ("QPS", f"{stats.get('qps', 0):.2f}"),
            ("RPS", f"{stats.get('rps', 0):.2f}"),
        ]
        for label, value in stats_items:
            pdf.cell(60, 7, f"  {label}:", border=0)
            pdf.cell(0, 7, value, ln=True)

        pdf.ln(5)

        per_method = report_data.get("per_method_stats", {})
        if per_method:
            pdf.set_font(title_font, "B", 14)
            pdf.cell(0, 10, "Per-Method Statistics", ln=True)
            pdf.set_font(title_font, "B", 9)

            col_widths = [45, 25, 25, 30, 30, 30]
            headers = ["Endpoint", "Requests", "Failures", "Avg(ms)", "P95(ms)", "RPS"]
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 7, h, border=1)
            pdf.ln()

            pdf.set_font(body_font, "", 9)
            for endpoint, entry in per_method.items():
                pdf.cell(col_widths[0], 7, str(endpoint)[:25], border=1)
                pdf.cell(col_widths[1], 7, str(entry.get("num_requests", 0)), border=1)
                pdf.cell(col_widths[2], 7, str(entry.get("num_failures", 0)), border=1)
                pdf.cell(col_widths[3], 7, f"{entry.get('avg_response_time', 0):.2f}", border=1)
                pdf.cell(col_widths[4], 7, f"{entry.get('p95_response_time', 0):.2f}", border=1)
                pdf.cell(col_widths[5], 7, f"{entry.get('rps', 0):.2f}", border=1)
                pdf.ln()

        errors = report_data.get("errors", [])
        if errors:
            pdf.ln(5)
            pdf.set_font(title_font, "B", 14)
            pdf.cell(0, 10, "Errors", ln=True)
            pdf.set_font(body_font, "", 9)
            for err in errors[:20]:
                err_text = f"{err.get('method', '')} {err.get('name', '')}: {err.get('error', '')} ({err.get('occurrences', 0)} times)"
                pdf.multi_cell(0, 6, f"  - {err_text}")

        pdf.output(str(output_path))
        logger.info("PDF报告已生成，路径=%s", output_path)
        return output_path

    def export_excel(
        self,
        task_id: int,
        result_id: int | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """导出Excel报告

        使用openpyxl库生成Excel格式的测试报告，
        包含概览、统计数据、按接口统计、错误列表等多个工作表。

        Args:
            task_id: 任务ID
            result_id: 结果记录ID，为None时使用最新结果
            output_path: 输出文件路径，为None时自动生成

        Returns:
            生成的Excel文件路径

        Raises:
            ValueError: 任务或结果不存在
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        report_data = self._collect_report_data(task_id, result_id)

        if output_path is None:
            from config.settings import get_settings
            settings = get_settings()
            export_dir = settings.export_dir
            ensure_dir(export_dir)
            safe_name = report_data["task_info"].get("name", "unnamed").replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = export_dir / f"report_{safe_name}_{timestamp}.xlsx"

        output_path = Path(output_path)
        ensure_dir(output_path.parent)

        wb = Workbook()

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 概览工作表
        ws_overview = wb.active
        ws_overview.title = "Overview"
        ws_overview.column_dimensions["A"].width = 25
        ws_overview.column_dimensions["B"].width = 50

        task_info = report_data["task_info"]
        overview_data = [
            ("Task Name", task_info.get("name", "")),
            ("Task Type", task_info.get("type", "")),
            ("Method", task_info.get("method", "")),
            ("URL", task_info.get("url", "")),
            ("Users", task_info.get("users", "")),
            ("Spawn Rate", task_info.get("spawn_rate", "")),
            ("Run Time", task_info.get("run_time", "")),
            ("Timeout", task_info.get("timeout", "")),
            ("Start Time", report_data["stats"].get("start_time", "")),
            ("End Time", report_data["stats"].get("end_time", "")),
        ]

        for col_idx, header in enumerate(["Property", "Value"], 1):
            cell = ws_overview.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (prop, value) in enumerate(overview_data, 2):
            ws_overview.cell(row=row_idx, column=1, value=prop).border = thin_border
            ws_overview.cell(row=row_idx, column=2, value=str(value)).border = thin_border

        # 统计数据工作表
        ws_stats = wb.create_sheet("Statistics")
        ws_stats.column_dimensions["A"].width = 30
        ws_stats.column_dimensions["B"].width = 20

        stats = report_data["stats"]
        stats_data = [
            ("Total Requests", stats.get("total_requests", 0)),
            ("Success Count", stats.get("success_count", 0)),
            ("Fail Count", stats.get("fail_count", 0)),
            ("Fail Rate", f"{stats.get('fail_rate', 0):.2%}"),
            ("Avg Response Time (ms)", round(stats.get("avg_response_time", 0), 2)),
            ("Min Response Time (ms)", round(stats.get("min_response_time", 0), 2)),
            ("Max Response Time (ms)", round(stats.get("max_response_time", 0), 2)),
            ("P95 Response Time (ms)", round(stats.get("p95_response_time", 0), 2)),
            ("QPS", round(stats.get("qps", 0), 2)),
            ("RPS", round(stats.get("rps", 0), 2)),
        ]

        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws_stats.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (metric, value) in enumerate(stats_data, 2):
            ws_stats.cell(row=row_idx, column=1, value=metric).border = thin_border
            ws_stats.cell(row=row_idx, column=2, value=str(value)).border = thin_border

        # 按接口统计工作表
        per_method = report_data.get("per_method_stats", {})
        if per_method:
            ws_method = wb.create_sheet("Per-Method Stats")
            method_headers = [
                "Endpoint", "Requests", "Failures",
                "Avg RT (ms)", "Min RT (ms)", "Max RT (ms)",
                "P95 RT (ms)", "RPS",
            ]
            for col_idx, header in enumerate(method_headers, 1):
                cell = ws_method.cell(row=1, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, (endpoint, entry) in enumerate(per_method.items(), 2):
                ws_method.cell(row=row_idx, column=1, value=endpoint).border = thin_border
                ws_method.cell(row=row_idx, column=2, value=entry.get("num_requests", 0)).border = thin_border
                ws_method.cell(row=row_idx, column=3, value=entry.get("num_failures", 0)).border = thin_border
                ws_method.cell(row=row_idx, column=4, value=round(entry.get("avg_response_time", 0), 2)).border = thin_border
                ws_method.cell(row=row_idx, column=5, value=round(entry.get("min_response_time", 0), 2)).border = thin_border
                ws_method.cell(row=row_idx, column=6, value=round(entry.get("max_response_time", 0), 2)).border = thin_border
                ws_method.cell(row=row_idx, column=7, value=round(entry.get("p95_response_time", 0), 2)).border = thin_border
                ws_method.cell(row=row_idx, column=8, value=round(entry.get("rps", 0), 2)).border = thin_border

            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                ws_method.column_dimensions[col_letter].width = 18

        # 错误列表工作表
        errors = report_data.get("errors", [])
        if errors:
            ws_errors = wb.create_sheet("Errors")
            error_headers = ["Method", "Name", "Error", "Occurrences"]
            for col_idx, header in enumerate(error_headers, 1):
                cell = ws_errors.cell(row=1, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, err in enumerate(errors, 2):
                ws_errors.cell(row=row_idx, column=1, value=err.get("method", "")).border = thin_border
                ws_errors.cell(row=row_idx, column=2, value=err.get("name", "")).border = thin_border
                ws_errors.cell(row=row_idx, column=3, value=err.get("error", "")).border = thin_border
                ws_errors.cell(row=row_idx, column=4, value=err.get("occurrences", 0)).border = thin_border

            for col_letter in ["A", "B", "C", "D"]:
                ws_errors.column_dimensions[col_letter].width = 30

        wb.save(str(output_path))
        logger.info("Excel报告已生成，路径=%s", output_path)
        return output_path

    def export_png_chart(
        self,
        task_id: int,
        result_id: int | None = None,
        output_path: str | Path | None = None,
        chart_type: str = "response_time",
    ) -> Path:
        """导出PNG图表

        使用matplotlib生成图表并保存为PNG格式。
        支持响应时间分布图和状态码分布图。

        Args:
            task_id: 任务ID
            result_id: 结果记录ID，为None时使用最新结果
            output_path: 输出文件路径，为None时自动生成
            chart_type: 图表类型，"response_time" 或 "status_code"

        Returns:
            生成的图表文件路径

        Raises:
            ValueError: 任务或结果不存在，或不支持的图表类型
        """
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm

        report_data = self._collect_report_data(task_id, result_id)

        if output_path is None:
            from config.settings import get_settings
            settings = get_settings()
            export_dir = settings.export_dir
            ensure_dir(export_dir)
            safe_name = report_data["task_info"].get("name", "unnamed").replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = export_dir / f"chart_{chart_type}_{safe_name}_{timestamp}.png"

        output_path = Path(output_path)
        ensure_dir(output_path.parent)

        stats = report_data["stats"]
        per_method = report_data.get("per_method_stats", {})

        if chart_type == "response_time":
            self._draw_response_time_chart(per_method, stats, output_path)
        elif chart_type == "status_code":
            self._draw_status_code_chart(stats, output_path)
        else:
            raise ValueError(f"不支持的图表类型: {chart_type}，有效值: response_time, status_code")

        logger.info("PNG图表已生成，类型=%s，路径=%s", chart_type, output_path)
        return output_path

    def _draw_response_time_chart(
        self,
        per_method: dict[str, Any],
        stats: dict[str, Any],
        output_path: Path,
    ) -> None:
        """绘制响应时间分布图

        Args:
            per_method: 按接口统计数据
            stats: 总体统计数据
            output_path: 输出路径
        """
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, axes = plt.subplots(1, 2, figsize=(14, 6))

        if per_method:
            endpoints = list(per_method.keys())
            avg_times = [per_method[e].get("avg_response_time", 0) for e in endpoints]
            p95_times = [per_method[e].get("p95_response_time", 0) for e in endpoints]

            short_labels = [e[:20] + "..." if len(e) > 20 else e for e in endpoints]

            x = range(len(short_labels))
            width = 0.35

            axes[0].bar([i - width / 2 for i in x], avg_times, width, label="Avg RT", color="#4472C4")
            axes[0].bar([i + width / 2 for i in x], p95_times, width, label="P95 RT", color="#ED7D31")
            axes[0].set_xlabel("Endpoint")
            axes[0].set_ylabel("Response Time (ms)")
            axes[0].set_title("Response Time by Endpoint")
            axes[0].set_xticks(list(x))
            axes[0].set_xticklabels(short_labels, rotation=45, ha="right", fontsize=8)
            axes[0].legend()
        else:
            axes[0].text(0.5, 0.5, "No per-method data", ha="center", va="center")
            axes[0].set_title("Response Time by Endpoint")

        summary_metrics = ["Avg", "Min", "Max", "P95"]
        summary_values = [
            stats.get("avg_response_time", 0),
            stats.get("min_response_time", 0),
            stats.get("max_response_time", 0),
            stats.get("p95_response_time", 0),
        ]
        colors = ["#4472C4", "#70AD47", "#FFC000", "#ED7D31"]

        bars = axes[1].bar(summary_metrics, summary_values, color=colors)
        axes[1].set_ylabel("Response Time (ms)")
        axes[1].set_title("Overall Response Time Summary")

        for bar, val in zip(bars, summary_values):
            axes[1].text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{val:.1f}",
                ha="center",
                va="bottom",
                fontsize=9,
            )

        plt.tight_layout()
        plt.savefig(str(output_path), dpi=150, bbox_inches="tight")
        plt.close(fig)

    def _draw_status_code_chart(
        self,
        stats: dict[str, Any],
        output_path: Path,
    ) -> None:
        """绘制状态码分布图

        Args:
            stats: 统计数据
            output_path: 输出路径
        """
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(8, 6))

        total_requests = stats.get("total_requests", 0)
        total_failures = stats.get("total_failures", 0)
        success_count = total_requests - total_failures

        if total_requests > 0:
            labels = ["Success", "Failure"]
            sizes = [success_count, total_failures]
            colors = ["#70AD47", "#FF0000"]
            explode = [0, 0.05] if total_failures > 0 else [0, 0]

            wedges, texts, autotexts = ax.pie(
                sizes,
                explode=explode,
                labels=labels,
                colors=colors,
                autopct="%1.1f%%",
                startangle=90,
            )
            for text in autotexts:
                text.set_fontsize(10)
        else:
            ax.text(0.5, 0.5, "No request data", ha="center", va="center")

        ax.set_title(f"Request Status Distribution (Total: {total_requests})")
        plt.tight_layout()
        plt.savefig(str(output_path), dpi=150, bbox_inches="tight")
        plt.close(fig)

    def _collect_report_data(
        self,
        task_id: int,
        result_id: int | None = None,
    ) -> dict[str, Any]:
        """收集报告所需的数据

        从数据库中查询任务信息和执行结果，
        整合为报告渲染所需的字典结构。

        Args:
            task_id: 任务ID
            result_id: 结果记录ID，为None时使用最新结果

        Returns:
            报告数据字典

        Raises:
            ValueError: 任务或结果不存在
        """
        task = self._db.get_task(task_id)
        if task is None:
            raise ValueError(f"任务不存在，ID: {task_id}")

        if result_id is not None:
            result = self._db.get_task_result(result_id)
        else:
            result = self._db.get_latest_result_by_task(task_id)

        if result is None:
            raise ValueError(f"任务结果不存在，任务ID: {task_id}")

        stats_json = result.get("stats_json", {})
        if isinstance(stats_json, str):
            try:
                stats_json = json.loads(stats_json)
            except (json.JSONDecodeError, TypeError):
                stats_json = {}

        per_method_stats = stats_json.get("requests_per_method", {})
        errors = stats_json.get("errors", [])

        stats = {
            "total_requests": result.get("total_requests", 0),
            "success_count": result.get("success_count", 0),
            "fail_count": result.get("fail_count", 0),
            "fail_rate": result.get("fail_rate", 0.0),
            "avg_response_time": result.get("avg_response_time", 0.0),
            "min_response_time": result.get("min_response_time", 0.0),
            "max_response_time": result.get("max_response_time", 0.0),
            "p95_response_time": result.get("p95_response_time", 0.0),
            "qps": result.get("qps", 0.0),
            "tps": result.get("tps", 0.0),
            "rps": result.get("rps", 0.0),
            "current_users": result.get("current_users", 0),
            "start_time": result.get("start_time", ""),
            "end_time": result.get("end_time", ""),
        }

        return {
            "task_info": task,
            "result_info": result,
            "stats": stats,
            "per_method_stats": per_method_stats,
            "errors": errors,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    _HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Performance Test Report - {{ task_info.name | default('Unnamed') }}</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f7fa; color: #333; line-height: 1.6; }
  .container { max-width: 1100px; margin: 0 auto; padding: 20px; }
  .header { background: linear-gradient(135deg, #4472C4, #2B5797); color: #fff; padding: 30px; border-radius: 8px; margin-bottom: 20px; }
  .header h1 { font-size: 24px; margin-bottom: 8px; }
  .header p { font-size: 14px; opacity: 0.85; }
  .card { background: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 24px; margin-bottom: 20px; }
  .card h2 { font-size: 18px; color: #2B5797; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #4472C4; }
  .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; }
  .stat-item { background: #f8f9fb; border-radius: 6px; padding: 16px; text-align: center; }
  .stat-item .label { font-size: 12px; color: #888; margin-bottom: 4px; }
  .stat-item .value { font-size: 22px; font-weight: 700; color: #2B5797; }
  .stat-item .value.danger { color: #e74c3c; }
  .stat-item .value.success { color: #27ae60; }
  table { width: 100%; border-collapse: collapse; margin-top: 12px; }
  th, td { padding: 10px 14px; text-align: left; border-bottom: 1px solid #eee; font-size: 13px; }
  th { background: #4472C4; color: #fff; font-weight: 600; }
  tr:nth-child(even) { background: #f8f9fb; }
  tr:hover { background: #eef2f7; }
  .overview-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; }
  .overview-item { display: flex; justify-content: space-between; padding: 8px 12px; border-bottom: 1px solid #f0f0f0; }
  .overview-item .key { color: #666; font-size: 13px; }
  .overview-item .val { font-weight: 600; font-size: 13px; }
  .error-list { max-height: 300px; overflow-y: auto; }
  .error-item { padding: 8px 12px; background: #fff5f5; border-left: 3px solid #e74c3c; margin-bottom: 6px; border-radius: 4px; font-size: 13px; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }
  .badge-success { background: #d4edda; color: #155724; }
  .badge-danger { background: #f8d7da; color: #721c24; }
  .chart-placeholder { text-align: center; padding: 40px; color: #aaa; }
  @media print { body { background: #fff; } .card { box-shadow: none; border: 1px solid #ddd; } }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>Performance Test Report</h1>
    <p>Task: {{ task_info.name | default('Unnamed') }} &nbsp;|&nbsp; Generated: {{ generated_at }}</p>
  </div>

  <div class="card">
    <h2>Task Overview</h2>
    <div class="overview-grid">
      <div class="overview-item"><span class="key">Task Name</span><span class="val">{{ task_info.name | default('-') }}</span></div>
      <div class="overview-item"><span class="key">Task Type</span><span class="val">{{ task_info.type | default('-') }}</span></div>
      <div class="overview-item"><span class="key">Method</span><span class="val">{{ task_info.method | default('-') }}</span></div>
      <div class="overview-item"><span class="key">URL</span><span class="val">{{ task_info.url | default('-') }}</span></div>
      <div class="overview-item"><span class="key">Users</span><span class="val">{{ task_info.users | default('-') }}</span></div>
      <div class="overview-item"><span class="key">Spawn Rate</span><span class="val">{{ task_info.spawn_rate | default('-') }}/s</span></div>
      <div class="overview-item"><span class="key">Run Time</span><span class="val">{{ task_info.run_time | default('-') }}</span></div>
      <div class="overview-item"><span class="key">Timeout</span><span class="val">{{ task_info.timeout | default('-') }}s</span></div>
      <div class="overview-item"><span class="key">Start Time</span><span class="val">{{ stats.start_time | default('-') }}</span></div>
      <div class="overview-item"><span class="key">End Time</span><span class="val">{{ stats.end_time | default('-') }}</span></div>
    </div>
  </div>

  <div class="card">
    <h2>Statistics Summary</h2>
    <div class="grid">
      <div class="stat-item">
        <div class="label">Total Requests</div>
        <div class="value">{{ stats.total_requests | default(0) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">Success Count</div>
        <div class="value success">{{ stats.success_count | default(0) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">Fail Count</div>
        <div class="value danger">{{ stats.fail_count | default(0) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">Fail Rate</div>
        <div class="value {{ 'danger' if stats.fail_rate > 0.01 else 'success' }}">{{ "%.2f%%" | format(stats.fail_rate * 100 if stats.fail_rate else 0) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">Avg Response Time</div>
        <div class="value">{{ "%.2f" | format(stats.avg_response_time | default(0)) }} ms</div>
      </div>
      <div class="stat-item">
        <div class="label">P95 Response Time</div>
        <div class="value">{{ "%.2f" | format(stats.p95_response_time | default(0)) }} ms</div>
      </div>
      <div class="stat-item">
        <div class="label">Min Response Time</div>
        <div class="value">{{ "%.2f" | format(stats.min_response_time | default(0)) }} ms</div>
      </div>
      <div class="stat-item">
        <div class="label">Max Response Time</div>
        <div class="value">{{ "%.2f" | format(stats.max_response_time | default(0)) }} ms</div>
      </div>
      <div class="stat-item">
        <div class="label">QPS</div>
        <div class="value">{{ "%.2f" | format(stats.qps | default(0)) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">RPS</div>
        <div class="value">{{ "%.2f" | format(stats.rps | default(0)) }}</div>
      </div>
    </div>
  </div>

  {% if per_method_stats %}
  <div class="card">
    <h2>Response Time Distribution</h2>
    <table>
      <thead>
        <tr>
          <th>Endpoint</th>
          <th>Requests</th>
          <th>Failures</th>
          <th>Avg RT (ms)</th>
          <th>Min RT (ms)</th>
          <th>Max RT (ms)</th>
          <th>P95 RT (ms)</th>
          <th>RPS</th>
        </tr>
      </thead>
      <tbody>
        {% for endpoint, entry in per_method_stats.items() %}
        <tr>
          <td>{{ endpoint }}</td>
          <td>{{ entry.num_requests | default(0) }}</td>
          <td><span class="badge {{ 'badge-danger' if entry.num_failures > 0 else 'badge-success' }}">{{ entry.num_failures | default(0) }}</span></td>
          <td>{{ "%.2f" | format(entry.avg_response_time | default(0)) }}</td>
          <td>{{ "%.2f" | format(entry.min_response_time | default(0)) }}</td>
          <td>{{ "%.2f" | format(entry.max_response_time | default(0)) }}</td>
          <td>{{ "%.2f" | format(entry.p95_response_time | default(0)) }}</td>
          <td>{{ "%.2f" | format(entry.rps | default(0)) }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}

  <div class="card">
    <h2>Status Code Distribution</h2>
    <div class="grid" style="grid-template-columns: repeat(2, 1fr);">
      <div class="stat-item">
        <div class="label">Success (2xx/3xx)</div>
        <div class="value success">{{ stats.success_count | default(0) }}</div>
      </div>
      <div class="stat-item">
        <div class="label">Failure (4xx/5xx)</div>
        <div class="value danger">{{ stats.fail_count | default(0) }}</div>
      </div>
    </div>
  </div>

  {% if errors %}
  <div class="card">
    <h2>Errors ({{ errors | length }})</h2>
    <div class="error-list">
      {% for err in errors %}
      <div class="error-item">
        <strong>{{ err.method | default('') }} {{ err.name | default('') }}</strong>: {{ err.error | default('') }}
        <span style="float:right; color:#999;">{{ err.occurrences | default(0) }} times</span>
      </div>
      {% endfor %}
    </div>
  </div>
  {% endif %}

</div>
</body>
</html>
"""
